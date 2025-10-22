import cv2
import numpy as np
from openpyxl.workbook import Workbook

import helper
import config

import tkinter as tk
from tkinter import filedialog
from PIL import Image, ImageTk
import os

id=[]
res=[]
def processImages(filepath):


    img = cv2.imread(filepath)
    filename = os.path.basename(filepath)
    print("Filename:", filename)
    filename_without_extension = filename.split('.')[0]
    print("Filename without extension:", filename_without_extension)
    id.append(filename_without_extension)

    # Get configuration values
    width = config.IMAGE_WIDTH
    height = config.IMAGE_HEIGHT
    questions = config.NUM_QUESTIONS
    choices = config.NUM_CHOICES
    ans = config.ANSWER_KEY

    # RESIZING
    img = cv2.resize(img, (width, height))
    cv2.imshow("Input image", img)
    cv2.waitKey(0)

    imgGray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    cv2.imshow("Grayscaled image", imgGray)
    cv2.waitKey(0)

    # BLURRING - Optimized using OpenCV's built-in GaussianBlur
    # GaussianBlur is much faster than manual convolution
    # kernel size should be odd (5x5 is equivalent to sigma=1)
    imgBlur = cv2.GaussianBlur(imgGray, config.BLUR_KERNEL_SIZE, config.BLUR_SIGMA)
    cv2.imshow("Blurred image", imgBlur)
    cv2.waitKey(0)

    # EDGE DETECTION - Using OpenCV's optimized built-in Canny
    # cv2.Canny(image, threshold1, threshold2)
    imgCanny = cv2.Canny(imgBlur, config.CANNY_THRESHOLD1, config.CANNY_THRESHOLD2)
    cv2.imshow("Edge detected image", imgCanny)
    cv2.waitKey(0)





    # FINDING ALL CONTOURS
    contours = helper.get_edge_points(imgCanny)
    # print(contours)
    # print(len(contours)) 
    drawCnt = img.copy()
    for i in range(len(contours)):
        # print("Contours")
        # print(contours[i])
        # contour_points = np.array(contours[i], dtype=np.int32)
        helper.manual_draw_contours(drawCnt, contours[i], (0, 255, 0), 1)
        # print("done")
    cv2.imshow("Contours", drawCnt)
    cv2.waitKey(0)




    # FIND CORNERS - Get bounding box for each contour
    corner_points = []
    for contour in contours:
        corner_list = helper.find_corners(contour)
        corner_points.append(corner_list)
    # print(corner_points)

    drawCorners = img.copy()
    for corner in corner_points:
        helper.manual_draw_contours(drawCorners, corner, (0, 255, 0), 2)
    # cv2.imshow("Corners",drawCorners)
    # cv2.waitKey(0)

    # FIND AREA
    areas = []
    for corners in corner_points:
        rec_width1 = abs(corners[1][0] - corners[0][0])
        rec_width2 = abs(corners[3][0] - corners[2][0])
        rec_width = (rec_width2 + rec_width1) / 2
        rec_height1 = abs(corners[0][1] - corners[2][1])
        rec_height2 = abs(corners[1][1] - corners[3][1])
        rec_height = (rec_height2 + rec_height1) / 2
        area = rec_width * rec_height
        areas.append(area)
    print("Area")
    print(areas)

    sorted_area = sorted(areas, reverse=True)
    print(sorted_area)
    max_area = sorted_area[0]
    area_index = []
    for i in range(len(sorted_area)):
        for j in range(len(areas)):
            if (sorted_area[i] == areas[j]):
                area_index.append(j)
                break;
    # print(area_index)
    max_index = area_index[0]
    max_contour = img.copy()
    # 0 -> ans, 2 -> grade, 4 -> name
    helper.manual_draw_contours(max_contour, contours[area_index[0]], (0, 255, 0), 1)

    cv2.imshow("Answer section", max_contour)
    cv2.waitKey(0)

    ans_corner_points = corner_points[max_index]
    bl_x = ans_corner_points[2][0] + config.ANSWER_CORNER_OFFSET['bottom_left_x']
    bl_y = ans_corner_points[2][1] + config.ANSWER_CORNER_OFFSET['bottom_left_y']
    tr_x = ans_corner_points[1][0] + config.ANSWER_CORNER_OFFSET['top_right_x']
    tr_y = ans_corner_points[1][1] + config.ANSWER_CORNER_OFFSET['top_right_y']
    # anss = np.ones((row_end-row_start+1,col_end-col_start+1))
    wd = tr_x - bl_x
    ht = tr_y - bl_y
    x, y, w, h = bl_x, bl_y, wd, ht  # Example: x, y, width, height of the ROI
    roi = imgGray[y:y + h, x:x + w]
    ans_new_image = np.zeros_like(roi)  # Create a black image with the same size as roi
    ans_new_image[:, :] = roi
    # cv2.imshow('Birds eye view', ans_new_image)
    # cv2.waitKey(0)

    height, width = ans_new_image.shape[:2]
    # Define the crop region for the lower portion
    crop_height = int(height * config.CROP_RATIO)  # Calculate the crop height

    # Crop the lower portion of the image
    lower_portion = ans_new_image[-crop_height:, :]
    img_padded = cv2.copyMakeBorder(lower_portion, 
                                    config.BORDER_PADDING['top'], 
                                    config.BORDER_PADDING['bottom'],
                                    config.BORDER_PADDING['left'], 
                                    config.BORDER_PADDING['right'], 
                                    cv2.BORDER_CONSTANT)
    cv2.imshow("Birds eye view", img_padded)
    # print(img_padded.shape)
    cv2.waitKey(0)

    # Optimized thresholding - no need to pre-create imgThres
    imgThres = helper.thresholdImage(img_padded, None, config.THRESHOLD_VALUE)
    cv2.imshow("Thresholded image", imgThres)
    cv2.waitKey(0)

    boxes = helper.splitBoxes(imgThres)
    # for i in range(25):
    #     cv2.imshow("Test box", boxes[i])
    #     cv2.waitKey(0)

    # GETTING NON ZERO PIXEL VALUES OF EACH BOX
    myPixelVal = np.zeros((questions, choices))
    countC = 0
    countR = 0

    for image in boxes:
        totalPixels = helper.countNonZeroPixel(image)
        # print("pixel count")
        # print(totalPixels)
        # cv2.imshow("Test box", image)
        # cv2.waitKey(0)

        myPixelVal[countR][countC] = totalPixels
        countC += 1
        if (countC == choices):
            countR += 1
            countC = 0
    # print("Count pixel value")
    # print(myPixelVal)

    # FINDING INDEXES OF THE MARKINGS
    myIndex = []
    for x in range(0, questions):
        arr = myPixelVal[x]
        myIndexVal = np.where(arr == np.amax(arr))
        # print(myIndexVal[0])
        myIndex.append(myIndexVal[0][0])
    # print(myIndex)

    # GRADING
    gradings = []
    for x in range(0, questions):
        if (ans[x] == myIndex[x]):
            gradings.append(1)
        else:
            gradings.append(0)
    score = (sum(gradings) / questions) * 100
    res.append(score)
    # print(score)

    # DISPLAYING ANSWERS
    imgResult = img_padded.copy()
    imgResult = helper.showAnswers(imgResult, myIndex, questions, ans, choices, gradings)
    cv2.imshow("Answers", imgResult)
    cv2.waitKey(0)

    # DISPLAY GRADING
    grade_index = area_index[2]
    grade_contour = img.copy()
    # 0 -> ans, 2 -> grade, 4 -> name
    # helper.manual_draw_contours(grade_contour, contours[grade_index], (0, 255, 0), 1)
    # cv2.imshow("Grade contour",grade_contour)
    # cv2.waitKey(0)

    grade_corner_points = corner_points[grade_index]
    # print("grade corner points")
    # print(grade_corner_points)
    grade_bl_x = grade_corner_points[2][0] + config.GRADE_CORNER_OFFSET['bottom_left_x']
    grade_bl_y = grade_corner_points[2][1] + config.GRADE_CORNER_OFFSET['bottom_left_y']
    grade_tr_x = grade_corner_points[1][0] + config.GRADE_CORNER_OFFSET['top_right_x']
    grade_tr_y = grade_corner_points[1][1] + config.GRADE_CORNER_OFFSET['top_right_y']
    # anss = np.ones((row_end-row_start+1,col_end-col_start+1))
    grade_wd = grade_tr_x - grade_bl_x
    grade_ht = grade_tr_y - grade_bl_y
    x, y, w, h = grade_bl_x, grade_bl_y, grade_wd, grade_ht  # Example: x, y, width, height of the ROI
    grade_roi = imgGray[y:y + h, x:x + w]
    grade_new_image = np.zeros_like(grade_roi)  # Create a black image with the same size as roi
    grade_new_image[:, :] = grade_roi
    helper.manual_draw_contours(grade_contour, contours[grade_index], (0, 255, 0), 1)
    cv2.imshow('Grading section', grade_contour)
    cv2.waitKey(0)

    imgGrading = grade_contour.copy()
    cv2.putText(imgGrading, str(int(score)) + "%", 
                config.GRADE_TEXT_POSITION, 
                config.GRADE_TEXT_FONT, 
                config.GRADE_TEXT_SCALE, 
                config.GRADE_TEXT_COLOR, 
                config.GRADE_TEXT_THICKNESS)
    cv2.imshow("Grading", imgGrading)
    cv2.waitKey(0)









    cv2.destroyAllWindows()

def show_result_summary(id_list, score_list):
    """Display comprehensive result summary in console"""
    
    print("\n" + "="*70)
    print(" "*20 + "📊 RESULT SUMMARY 📊")
    print("="*70)
    
    total_students = len(score_list)
    
    if total_students == 0:
        print("No results to display.")
        return
    
    # Calculate statistics
    average_score = sum(score_list) / total_students
    highest_score = max(score_list)
    lowest_score = min(score_list)
    
    # Pass/Fail analysis
    pass_mark = config.PASS_MARK
    passed_students = sum(1 for score in score_list if score >= pass_mark)
    failed_students = total_students - passed_students
    pass_rate = (passed_students / total_students) * 100
    
    # Grade distribution
    grade_a = sum(1 for score in score_list if score >= config.GRADE_BOUNDARIES['A'])
    grade_b = sum(1 for score in score_list if config.GRADE_BOUNDARIES['B'] <= score < config.GRADE_BOUNDARIES['A'])
    grade_c = sum(1 for score in score_list if config.GRADE_BOUNDARIES['C'] <= score < config.GRADE_BOUNDARIES['B'])
    grade_f = sum(1 for score in score_list if score < config.GRADE_BOUNDARIES['C'])
    
    # Display overall statistics
    print(f"\n📈 OVERALL STATISTICS:")
    print(f"   Total Students     : {total_students}")
    print(f"   Average Score      : {average_score:.2f}%")
    print(f"   Highest Score      : {highest_score:.2f}%")
    print(f"   Lowest Score       : {lowest_score:.2f}%")
    print(f"   Pass Rate          : {pass_rate:.2f}% ({passed_students}/{total_students})")
    
    print(f"\n📊 GRADE DISTRIBUTION:")
    print(f"   Grade A (80-100%)  : {grade_a} students ({(grade_a/total_students)*100:.1f}%)")
    print(f"   Grade B (60-79%)   : {grade_b} students ({(grade_b/total_students)*100:.1f}%)")
    print(f"   Grade C (40-59%)   : {grade_c} students ({(grade_c/total_students)*100:.1f}%)")
    print(f"   Grade F (0-39%)    : {grade_f} students ({(grade_f/total_students)*100:.1f}%)")
    
    # Display individual results
    print(f"\n📝 INDIVIDUAL RESULTS:")
    print("-"*70)
    print(f"{'No.':<6} {'Student ID':<25} {'Score':<12} {'Grade':<10} {'Status'}")
    print("-"*70)
    
    for i, (student_id, score) in enumerate(zip(id_list, score_list), 1):
        # Determine grade
        if score >= config.GRADE_BOUNDARIES['A']:
            grade = "A"
        elif score >= config.GRADE_BOUNDARIES['B']:
            grade = "B"
        elif score >= config.GRADE_BOUNDARIES['C']:
            grade = "C"
        else:
            grade = "F"
        
        # Determine status
        status = "✓ PASS" if score >= pass_mark else "✗ FAIL"
        
        print(f"{i:<6} {student_id:<25} {score:>6.2f}%     {grade:<10} {status}")
    
    print("-"*70)
    
    # Performance insights
    print(f"\n💡 INSIGHTS:")
    if pass_rate >= 80:
        print(f"   🌟 Excellent performance! {pass_rate:.1f}% pass rate.")
    elif pass_rate >= 60:
        print(f"   ✓ Good performance with {pass_rate:.1f}% pass rate.")
    elif pass_rate >= 40:
        print(f"   ⚠ Average performance. {pass_rate:.1f}% pass rate - needs improvement.")
    else:
        print(f"   ⚠ Poor performance. Only {pass_rate:.1f}% pass rate - significant improvement needed.")
    
    if highest_score == 100:
        print(f"   🏆 Perfect score achieved!")
    
    if average_score < 50:
        print(f"   📚 Class average is below 50%. Consider reviewing difficult topics.")
    
    print("\n" + "="*70 + "\n")

def print_result():
    global uploaded_images
    if uploaded_images:

        # result_text = "Uploaded images:\n"
        for file_path, _ in uploaded_images:
            print(file_path)
            processImages(file_path)

    else:
        result_label.config(text="No images uploaded yet.")
    
    # Display summary statistics if multiple students were processed
    if len(id) > 0:
        show_result_summary(id, res)
    
    # Save to Excel
    save_to_excel(id,res)

    if print_result:
        result_text = "Results saved successfully!"
        result_label.config(text=result_text)



def save_to_excel(id_array, score_array):
    wb = Workbook()
    ws = wb.active
    ws.title = "ID Scores"

    # Write headers
    ws['A1'] = "No."
    ws['B1'] = "Student ID"
    ws['C1'] = "Score (%)"
    ws['D1'] = "Grade"
    ws['E1'] = "Status"

    # Calculate statistics for summary
    total_students = len(score_array)
    pass_mark = config.PASS_MARK
    
    # Write individual student data
    for idx, (id_val, score_val) in enumerate(zip(id_array, score_array), start=2):
        # Determine grade
        if score_val >= config.GRADE_BOUNDARIES['A']:
            grade = "A"
        elif score_val >= config.GRADE_BOUNDARIES['B']:
            grade = "B"
        elif score_val >= config.GRADE_BOUNDARIES['C']:
            grade = "C"
        else:
            grade = "F"
        
        # Determine status
        status = "PASS" if score_val >= pass_mark else "FAIL"
        
        ws[f'A{idx}'] = idx - 1  # Student number
        ws[f'B{idx}'] = id_val
        ws[f'C{idx}'] = round(score_val, 2)
        ws[f'D{idx}'] = grade
        ws[f'E{idx}'] = status
    
    # Add summary statistics at the bottom
    summary_row = len(score_array) + 3
    
    ws[f'A{summary_row}'] = "SUMMARY STATISTICS"
    ws[f'A{summary_row}'].font = ws[f'A{summary_row}'].font.copy(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Students:"
    ws[f'B{summary_row}'] = total_students
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average Score:"
    ws[f'B{summary_row}'] = round(sum(score_array) / total_students, 2) if total_students > 0 else 0
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Highest Score:"
    ws[f'B{summary_row}'] = round(max(score_array), 2) if score_array else 0
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Lowest Score:"
    ws[f'B{summary_row}'] = round(min(score_array), 2) if score_array else 0
    
    summary_row += 1
    passed_students = sum(1 for score in score_array if score >= pass_mark)
    ws[f'A{summary_row}'] = "Pass Rate:"
    ws[f'B{summary_row}'] = f"{(passed_students / total_students * 100):.2f}%" if total_students > 0 else "0%"
    
    summary_row += 2
    ws[f'A{summary_row}'] = "GRADE DISTRIBUTION"
    ws[f'A{summary_row}'].font = ws[f'A{summary_row}'].font.copy(bold=True)
    
    summary_row += 1
    grade_a = sum(1 for score in score_array if score >= config.GRADE_BOUNDARIES['A'])
    ws[f'A{summary_row}'] = f"Grade A ({config.GRADE_BOUNDARIES['A']}-100%):"
    ws[f'B{summary_row}'] = grade_a
    ws[f'C{summary_row}'] = f"{(grade_a/total_students)*100:.1f}%" if total_students > 0 else "0%"
    
    summary_row += 1
    grade_b = sum(1 for score in score_array if config.GRADE_BOUNDARIES['B'] <= score < config.GRADE_BOUNDARIES['A'])
    ws[f'A{summary_row}'] = f"Grade B ({config.GRADE_BOUNDARIES['B']}-{config.GRADE_BOUNDARIES['A']-1}%):"
    ws[f'B{summary_row}'] = grade_b
    ws[f'C{summary_row}'] = f"{(grade_b/total_students)*100:.1f}%" if total_students > 0 else "0%"
    
    summary_row += 1
    grade_c = sum(1 for score in score_array if config.GRADE_BOUNDARIES['C'] <= score < config.GRADE_BOUNDARIES['B'])
    ws[f'A{summary_row}'] = f"Grade C ({config.GRADE_BOUNDARIES['C']}-{config.GRADE_BOUNDARIES['B']-1}%):"
    ws[f'B{summary_row}'] = grade_c
    ws[f'C{summary_row}'] = f"{(grade_c/total_students)*100:.1f}%" if total_students > 0 else "0%"
    
    summary_row += 1
    grade_f = sum(1 for score in score_array if score < config.GRADE_BOUNDARIES['C'])
    ws[f'A{summary_row}'] = f"Grade F (0-{config.GRADE_BOUNDARIES['C']-1}%):"
    ws[f'B{summary_row}'] = grade_f
    ws[f'C{summary_row}'] = f"{(grade_f/total_students)*100:.1f}%" if total_students > 0 else "0%"

    # Adjust column widths for better readability
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10

    # Save the workbook
    excel_filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Result", "*.xlsx")],
                                                  title="Save Excel file")
    if excel_filename:
        wb.save(excel_filename)
        print(f"Excel file saved to: {excel_filename}")

def rounded_button(parent, text, command):
    button = tk.Button(parent, text=text, command=command, bg=config.GUI_BUTTON_COLOR, fg=config.GUI_TEXT_COLOR, 
                       padx=15, pady=6, relief="flat", font=("Helvetica", 12), width=20, height=2)
    button.config(borderwidth=0, highlightthickness=0, bd=0)
    button.pack(pady=10)
    return button


def upload_images():
    global uploaded_images
    global thumbnail_images
    file_paths = filedialog.askopenfilenames(
        filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.gif;*.bmp")],
        title="Choose images to upload"
    )
    for file_path in file_paths:
        image = Image.open(file_path)
        image.thumbnail(config.THUMBNAIL_SIZE)  # Resize the image to a thumbnail
        uploaded_images.append((file_path, image))
        thumbnail_images.append(ImageTk.PhotoImage(image))

    if uploaded_images:
        result_text = "Images uploaded successfully!"
        result_label.config(text=result_text)
    else:
        result_label.config(text="No images uploaded yet.")
    show_images()





def show_images():
    # Clear previous thumbnails
    for widget in image_frame.winfo_children():
        widget.destroy()

    # Display thumbnails
    for img_tk in thumbnail_images:
        label = tk.Label(image_frame, image=img_tk, bg=config.GUI_BG_COLOR)
        label.pack(side=tk.LEFT, padx=10, pady=10)

    # Update the scroll region to include all widgets in the image_frame
    image_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

def on_mouse_wheel(event):
    if event.delta:
        canvas.xview_scroll(int(-1*(event.delta/120)), "units")  # For Windows and MacOS
    else:
        canvas.xview_scroll(-1 if event.num == 5 else 1, "units")  # For Linux

if __name__ == "__main__":
    uploaded_images = []
    thumbnail_images = []

    root = tk.Tk()
    root.title("Optical Mark Recognition")

    root.geometry(f"{config.GUI_WIDTH}x{config.GUI_HEIGHT}")  # Set initial window size

    # Set background color for the root window
    root.configure(bg=config.GUI_BG_COLOR)

    # Create a frame to contain all widgets with the same background color
    frame = tk.Frame(root, bg=config.GUI_BG_COLOR)
    frame.pack(fill=tk.BOTH, expand=True)

    upload_button = rounded_button(frame, "Upload Images", upload_images)
    upload_button.pack(pady=20, anchor=tk.CENTER)

    print_button = rounded_button(frame, "Print Result", print_result)
    print_button.pack(pady=10, anchor=tk.CENTER)

    result_label = tk.Label(frame, text="Please upload images", wraplength=500, bg=config.GUI_BG_COLOR, padx=10, pady=10,
                            font=("Helvetica", 14), fg=config.GUI_TEXT_COLOR)
    result_label.pack(pady=20, fill=tk.BOTH, expand=True)

    # Create a scrollable frame to display images
    canvas = tk.Canvas(frame, bg=config.GUI_BG_COLOR)
    scrollbar = tk.Scrollbar(frame, orient=tk.HORIZONTAL, command=canvas.xview)
    canvas.configure(xscrollcommand=scrollbar.set)

    scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
    canvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

    # Create a frame within the canvas to hold the images
    image_frame = tk.Frame(canvas, bg=config.GUI_BG_COLOR)
    canvas.create_window((0, 0), window=image_frame, anchor="nw")

    # Bind the frame to the canvas, and configure the canvas to adjust the scroll region
    image_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    # Bind mouse wheel to the canvas for scrolling
    canvas.bind_all("<MouseWheel>", on_mouse_wheel)  # Windows and MacOS
    canvas.bind_all("<Button-4>", on_mouse_wheel)  # Linux (scroll up)
    canvas.bind_all("<Button-5>", on_mouse_wheel)  # Linux (scroll down)

    root.mainloop()